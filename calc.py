# install OpenPyXl,streamlit
import streamlit as st
from openpyxl import load_workbook
import datetime


def chose_cell(filepath, quarter, cell, user_input):
    wb = load_workbook(filepath)
    ws = wb[quarter]
    try:
        if ws[cell].value is not None:
            ws[cell] = float(ws[cell].value) + float(user_input)
        else:
            ws[cell].value = 0
            wb.save(filepath)
    except ValueError:
        st.warning('Please enter and numeric value')
    wb.save(filepath)
    st.info(f'Stan wydatków na jedzenie w {current_time.strftime("%B")} wynosi {ws[cell].value}', icon="ℹ️")


# load in your workbook
FilePath = r'D:\Finanse\Finansy zycia.xlsx'
current_time = datetime.date.today()
st.title("Kalkulator wydatków na jedzenie")

inp = st.text_input(label="Input", placeholder="Dodaj nowy paragon...",
                    key='receipt')

if inp != '':
    st.error("Czy na pewno chcesz wprowadzić paragon?")
    if st.button("Nie"):
        exit()
    elif st.button("Tak"):
        if current_time.strftime('%B') is 'November':
            chose_cell(FilePath, 'Wydatki', 'B1', inp)
        elif current_time.strftime('%B') is "December":
            chose_cell(FilePath, 'Wydatki', 'B2', inp)
        elif current_time.strftime('%B') is "January":
            chose_cell(FilePath, 'Wydatki', 'B3', inp)
        elif current_time.strftime('%B') is "February":
            chose_cell(FilePath, 'Wydatki', 'B4', inp)
        elif current_time.strftime('%B') is "March":
            chose_cell(FilePath, 'Wydatki', 'B5', inp)
        elif current_time.strftime('%B') is "April":
            chose_cell(FilePath, 'Wydatki', 'B6', inp)
        elif current_time.strftime('%B') is "May":
            chose_cell(FilePath, 'Wydatki', 'B7', inp)
        elif current_time.strftime('%B') is "June":
            chose_cell(FilePath, 'Wydatki', 'B8', inp)
        elif current_time.strftime('%B') is "July":
            chose_cell(FilePath, 'Wydatki', 'B9', inp)
        elif current_time.strftime('%B') is "August":
            chose_cell(FilePath, 'Wydatki', 'B10', inp)
        elif current_time.strftime('%B') is "September":
            chose_cell(FilePath, 'Wydatki', 'B11', inp)
        elif current_time.strftime('%B') is "October":
            chose_cell(FilePath, 'Wydatki', 'B12', inp)
else:
    exit()
